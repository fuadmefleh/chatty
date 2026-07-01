"""Walmart order PDF and XLSX parser with SQLite database storage."""

import sqlite3
import re
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
import PyPDF2
import openpyxl


class WalmartOrderDB:
    """SQLite database manager for Walmart orders."""
    
    def __init__(self, db_path: str = "data/walmart/walmart_orders.db"):
        """Initialize database connection."""
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = None
        self.cursor = None
        self._connect()
        self._create_tables()
    
    def _connect(self):
        """Connect to the database."""
        self.conn = sqlite3.connect(str(self.db_path))
        self.cursor = self.conn.cursor()
    
    def _create_tables(self):
        """Create database tables if they don't exist."""
        # Orders table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                order_id TEXT PRIMARY KEY,
                order_date TEXT,
                total_amount REAL,
                subtotal REAL,
                tax REAL,
                shipping REAL,
                pdf_filename TEXT,
                parsed_date TEXT,
                raw_text TEXT
            )
        """)
        
        # Order items table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id TEXT,
                item_name TEXT,
                quantity INTEGER,
                unit_price REAL,
                total_price REAL,
                category TEXT,
                delivery_status TEXT,
                FOREIGN KEY (order_id) REFERENCES orders (order_id)
            )
        """)
        
        self.conn.commit()
        
        # Add category column if it doesn't exist (for existing databases)
        try:
            self.cursor.execute("ALTER TABLE order_items ADD COLUMN category TEXT")
            self.conn.commit()
        except sqlite3.OperationalError:
            # Column already exists
            pass
        
        # Add delivery_status column to order_items if it doesn't exist (for existing databases)
        try:
            self.cursor.execute("ALTER TABLE order_items ADD COLUMN delivery_status TEXT")
            self.conn.commit()
        except sqlite3.OperationalError:
            # Column already exists
            pass
    
    @staticmethod
    def categorize_item(item_name: str) -> str:
        """Categorize an item based on its name.
        
        Categories:
        - Food: Groceries, snacks, frozen food, fresh produce, meat, dairy
        - Beverages: Drinks, milk, juice, water, coffee
        - Household: Cleaning products, paper products, trash bags
        - Kitchen Supplies: Cookware, utensils, disposable items, foil
        - Personal Care: Toiletries, hygiene products
        - Pet Supplies: Pet food, pet care items
        - Health: Vitamins, medications, first aid
        - Baby & Kids: Baby products, diapers, kid-specific items
        - Other: Everything else
        """
        item_lower = item_name.lower()
        
        # Electronics - Check FIRST for Apple products and other electronics
        # Apple products must be checked before food to avoid 'apple' matching
        electronics_keywords = [
            'imac', 'macbook', 'ipad', 'iphone', 'airpods', 'apple watch',
            'mac mini', 'apple tv', 'apple pencil', 'apple m1', 'apple m2', 'apple m3', 'apple m4',
            'laptop', 'computer', 'tablet', 'smartphone', 'desktop',
            'cable', 'charger', 'adapter', 'usb', 'hdmi', 'bluetooth',
            'headphone', 'earbud', 'speaker', 'mouse', 'keyboard',
            'phone case', 'screen protector', 'battery', 'power bank',
            'monitor', 'webcam', 'carplay', 'playstation', 'xbox', 'nintendo'
        ]
        
        if any(keyword in item_lower for keyword in electronics_keywords):
            return 'Electronics'
        
        # High priority food indicators - check these AFTER electronics
        high_priority_food = [
            'fresh ', 'ice cream', 'frozen', 'refrigerated', 'yogurt', 'cheese',
            'bagel bites', 'waffle', 'cereal', 'sour cream', 'cream cheese',
            'cheddar', 'mozzarella', 'parmesan', 'gouda', 'brie',
            'apple', 'banana', 'orange', 'strawberr', 'berry', 'grape', 'melon', 'cantaloupe',
            'avocado', 'potato', 'carrot', 'onion', 'tomato', 'cucumber', 'zucchini',
            'lettuce', 'salad', 'broccoli', 'pepper', 'celery', 'mushroom',
            'chicken', 'beef', 'pork', 'meat', 'bacon', 'sausage', 'turkey', 'ham',
            'pizza', 'pasta', 'bread', 'cookie', 'cracker', 'chip',
            'chocolate', 'candy', 'dessert', 'cake', 'pie',
            'milk', 'egg', 'butter', 'bagel', 'barbecue sauce', 'bbq sauce',
            'sauce', 'ketchup', 'mustard', 'mayo', 'ranch'
        ]
        
        # Check high priority food first
        if any(keyword in item_lower for keyword in high_priority_food):
            # Exclude if it's actually a kitchen item or cleaning product
            if 'bowl' in item_lower or 'pan' in item_lower or 'foil' in item_lower:
                pass  # Continue to other checks
            elif 'wash' in item_lower or 'soap' in item_lower or 'cleaner' in item_lower:
                pass  # Continue to other checks
            else:
                return 'Food'
        
        # Beverages - check before general food to catch drinks
        beverage_keywords = [
            'water', 'juice', 'soda', 'pop', 'drink', 'beverage',
            'coffee', 'tea', 'lemonade', 'punch',
            'sports drink', 'energy drink', 'gatorade', 'powerade'
        ]
        
        if any(keyword in item_lower for keyword in beverage_keywords):
            # Exclude "milk" from beverages - it's food/dairy unless explicitly "milk drink"
            if 'milk' in item_lower and 'chocolate' not in item_lower:
                return 'Food'
            # Exclude body wash type products
            if 'wash' in item_lower or 'soap' in item_lower:
                pass  # Continue to other checks
            else:
                return 'Beverages'
        
        # Baby & Kids - check early for kid-specific items
        baby_keywords = [
            'diaper', 'infant', 'toddler', 'formula',
            'bluey', 'carter\'s', 'garanimals', 'child of mine',
            'sippy', 'bottle', 'pacifier', 'newborn', 'baby clothes',
            'baby outfit', 'onesie'
        ]
        
        # Exclude baby carrots/food - they're food
        if any(keyword in item_lower for keyword in baby_keywords):
            if 'baby-cut' in item_lower or 'baby carrot' in item_lower or 'baby peeled' in item_lower:
                return 'Food'
            # Exclude health items like hydrocortisone
            if 'cream' in item_lower and ('hydrocortisone' in item_lower or 'itch' in item_lower or 'eczema' in item_lower):
                return 'Health'
            # Exclude wipes if they're hand wipes
            if 'wipe' in item_lower and ('hand' in item_lower or 'antibacterial' in item_lower):
                return 'Personal Care'
            # Exclude food items that happen to have "baby" in the brand name
            if 'sauce' in item_lower or 'seasoning' in item_lower:
                return 'Food'
            return 'Baby & Kids'
        
        # Health - medical products
        health_keywords = [
            'vitamin', 'supplement', 'medicine', 'medication', 'pill',
            'bandage', 'first aid', 'aspirin', 'pain relief', 'antacid',
            'hydrocortisone', 'cortisone', 'anti itch', 'itch relief',
            'medicated', 'eczema', 'healing formula'
        ]
        
        if any(keyword in item_lower for keyword in health_keywords):
            return 'Health'
        
        # Personal Care
        personal_care_keywords = [
            'shampoo', 'conditioner', 'body wash', 'lotion', 'deodorant',
            'toothpaste', 'toothbrush', 'dental', 'razor', 'shave',
            'makeup', 'cosmetic', 'skincare', 'hair care', 'perfume', 'cologne',
            'hand wash', 'hand soap', 'antibacterial.*hand', 'moisturizer',
            'aveeno', 'dove', 'olay', 'nivea', 'gold bond'
        ]
        
        if any(keyword in item_lower for keyword in personal_care_keywords):
            return 'Personal Care'
        
        # Household cleaning & paper products
        household_keywords = [
            'detergent', 'bleach', 'disinfect', 'cleaner', 'cleaning',
            'paper towel', 'toilet paper', 'tissue',
            'trash bag', 'garbage bag', 'laundry', 'fabric softener',
            'dishwasher', 'dish soap', 'sponge', 'scrub', 'mop', 'broom',
            'cascade', 'tide', 'bounty', 'charmin', 'lysol', 'clorox',
            'dryer ball', 'wool.*ball'
        ]
        
        if any(keyword in item_lower for keyword in household_keywords):
            return 'Household'
        
        # Kitchen Supplies - non-food kitchen items
        kitchen_keywords = [
            'pan', 'pot', 'skillet', 'cookware',
            'foil', 'aluminum foil', 'plastic wrap', 'parchment',
            'ziploc', 'storage bag', 'food storage',
            'utensil', 'fork', 'spoon', 'knife', 'cutting board',
            'disposable plate', 'paper plate', 'paper cup', 'napkin'
        ]
        
        if any(keyword in item_lower for keyword in kitchen_keywords):
            return 'Kitchen Supplies'
        
        # Pet Supplies
        pet_keywords = [
            'dog', 'cat', 'pet', 'puppy', 'kitten', 'pet food',
            'pet treat', 'leash', 'collar', 'litter'
        ]
        
        # Exclude if it's actually food (e.g., "Jalapeno Peppers")
        if any(keyword in item_lower for keyword in pet_keywords):
            if 'pepper' in item_lower or 'jalapeno' in item_lower:
                return 'Food'
            return 'Pet Supplies'
        
        # General food keywords - catch anything remaining that's food
        food_keywords = [
            'snack', 'lunch', 'meal', 'dinner', 'breakfast', 'food',
            'soup', 'sauce', 'salsa', 'dressing', 'condiment', 'ketchup', 'mustard',
            'peanut butter', 'jelly', 'jam', 'honey', 'syrup', 'gravy',
            'rice', 'bean', 'taco', 'tortilla', 'burrito', 'quesadilla',
            'sandwich', 'wings', 'nugget', 'tender', 'strip',
            'spice', 'seasoning', 'salt', 'pepper', 'sugar', 'flour',
            'oil', 'vinegar', 'canned', 'can ',
            'pretzel', 'nut', 'trail mix', 'popcorn', 'bar',
            'wafer', 'melting wafer', 'baking', 'ghirardelli'
        ]
        
        if any(keyword in item_lower for keyword in food_keywords):
            return 'Food'
        
        return 'Other'
    
    def insert_order(self, order_data: Dict[str, Any], items: List[Dict[str, Any]]):
        """Insert an order and its items into the database."""
        try:
            # Insert order
            self.cursor.execute("""
                INSERT OR REPLACE INTO orders 
                (order_id, order_date, total_amount, subtotal, tax, shipping, 
                 pdf_filename, parsed_date, raw_text)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                order_data.get('order_id'),
                order_data.get('order_date'),
                order_data.get('total_amount'),
                order_data.get('subtotal'),
                order_data.get('tax'),
                order_data.get('shipping'),
                order_data.get('pdf_filename'),
                datetime.now().isoformat(),
                order_data.get('raw_text', '')
            ))
            
            # Delete existing items for this order (in case of re-parsing)
            self.cursor.execute("DELETE FROM order_items WHERE order_id = ?", 
                              (order_data.get('order_id'),))
            
            # Insert items
            for item in items:
                item_name = item.get('name')
                category = self.categorize_item(item_name)
                self.cursor.execute("""
                    INSERT INTO order_items 
                    (order_id, item_name, quantity, unit_price, total_price, category, delivery_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_data.get('order_id'),
                    item_name,
                    item.get('quantity', 1),
                    item.get('unit_price'),
                    item.get('total_price'),
                    category,
                    item.get('delivery_status')
                ))
            
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            raise Exception(f"Database error: {str(e)}")
    
    def get_order(self, order_id: str) -> Optional[Dict[str, Any]]:
        """Retrieve an order by ID with all items."""
        self.cursor.execute("""
            SELECT order_id, order_date, total_amount, subtotal, tax, shipping, 
                   pdf_filename, parsed_date
            FROM orders WHERE order_id = ?
        """, (order_id,))
        
        row = self.cursor.fetchone()
        if not row:
            return None
        
        order = {
            'order_id': row[0],
            'order_date': row[1],
            'total_amount': row[2],
            'subtotal': row[3],
            'tax': row[4],
            'shipping': row[5],
            'pdf_filename': row[6],
            'parsed_date': row[7]
        }
        
        # Get items
        self.cursor.execute("""
            SELECT item_name, quantity, unit_price, total_price, category, delivery_status
            FROM order_items WHERE order_id = ?
            ORDER BY id
        """, (order_id,))
        
        order['items'] = [
            {
                'name': row[0],
                'quantity': row[1],
                'unit_price': row[2],
                'total_price': row[3],
                'category': row[4],
                'delivery_status': row[5]
            }
            for row in self.cursor.fetchall()
        ]
        
        return order
    
    def get_all_orders(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Get all orders, most recent first."""
        self.cursor.execute("""
            SELECT order_id, order_date, total_amount, pdf_filename
            FROM orders 
            ORDER BY order_date DESC 
            LIMIT ?
        """, (limit,))
        
        return [
            {
                'order_id': row[0],
                'order_date': row[1],
                'total_amount': row[2],
                'pdf_filename': row[3]
            }
            for row in self.cursor.fetchall()
        ]
    
    def get_all_items(self, limit: int = 100) -> List[Dict[str, Any]]:
        """Get all items purchased across all orders."""
        self.cursor.execute("""
            SELECT oi.item_name, oi.quantity, oi.unit_price, oi.total_price,
                   o.order_id, o.order_date, oi.category
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.order_id
            ORDER BY o.order_date DESC, oi.id DESC
            LIMIT ?
        """, (limit,))
        
        return [
            {
                'name': row[0],
                'quantity': row[1],
                'unit_price': row[2],
                'total_price': row[3],
                'order_id': row[4],
                'order_date': row[5],
                'category': row[6]
            }
            for row in self.cursor.fetchall()
        ]
    
    def search_items(self, search_term: str) -> List[Dict[str, Any]]:
        """Search for items by name."""
        self.cursor.execute("""
            SELECT oi.item_name, SUM(oi.quantity) as total_qty, 
                   AVG(oi.unit_price) as avg_price, COUNT(DISTINCT o.order_id) as order_count,
                   oi.category
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.order_id
            WHERE oi.item_name LIKE ?
            GROUP BY oi.item_name, oi.category
            ORDER BY total_qty DESC
        """, (f'%{search_term}%',))
        
        return [
            {
                'name': row[0],
                'total_quantity': row[1],
                'avg_unit_price': round(row[2], 2),
                'times_purchased': row[3],
                'category': row[4]
            }
            for row in self.cursor.fetchall()
        ]
    
    def get_orders_by_date_range(self, start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """Get orders within a date range."""
        self.cursor.execute("""
            SELECT order_id, order_date, total_amount, pdf_filename
            FROM orders 
            WHERE order_date BETWEEN ? AND ?
            ORDER BY order_date DESC
        """, (start_date, end_date))
        
        return [
            {
                'order_id': row[0],
                'order_date': row[1],
                'total_amount': row[2],
                'pdf_filename': row[3]
            }
            for row in self.cursor.fetchall()
        ]
    
    def get_total_spent(self, start_date: Optional[str] = None, 
                       end_date: Optional[str] = None) -> float:
        """Get total amount spent in date range."""
        if start_date and end_date:
            self.cursor.execute("""
                SELECT SUM(total_amount) FROM orders
                WHERE order_date BETWEEN ? AND ?
            """, (start_date, end_date))
        else:
            self.cursor.execute("SELECT SUM(total_amount) FROM orders")
        
        result = self.cursor.fetchone()[0]
        return result if result else 0.0
    
    def get_spending_by_category(self, start_date: Optional[str] = None,
                                end_date: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get total spending broken down by category.
        
        Args:
            start_date: Optional start date filter
            end_date: Optional end date filter
            
        Returns:
            List of dicts with category, total_spent, item_count, and percentage
        """
        if start_date and end_date:
            self.cursor.execute("""
                SELECT oi.category, 
                       SUM(oi.total_price) as total_spent,
                       COUNT(*) as item_count,
                       SUM(oi.quantity) as total_quantity
                FROM order_items oi
                JOIN orders o ON oi.order_id = o.order_id
                WHERE o.order_date BETWEEN ? AND ?
                GROUP BY oi.category
                ORDER BY total_spent DESC
            """, (start_date, end_date))
        else:
            self.cursor.execute("""
                SELECT oi.category, 
                       SUM(oi.total_price) as total_spent,
                       COUNT(*) as item_count,
                       SUM(oi.quantity) as total_quantity
                FROM order_items oi
                GROUP BY oi.category
                ORDER BY total_spent DESC
            """)
        
        results = []
        total = 0.0
        rows = self.cursor.fetchall()
        
        # Calculate total for percentages
        for row in rows:
            total += row[1] if row[1] else 0.0
        
        # Build results with percentages
        for row in rows:
            spent = row[1] if row[1] else 0.0
            percentage = (spent / total * 100) if total > 0 else 0.0
            results.append({
                'category': row[0] if row[0] else 'Uncategorized',
                'total_spent': round(spent, 2),
                'item_count': row[2],
                'total_quantity': row[3],
                'percentage': round(percentage, 1)
            })
        
        return results
    
    def get_items_by_category(self, category: str, limit: int = 100) -> List[Dict[str, Any]]:
        """Get all items in a specific category.
        
        Args:
            category: Category name to filter by
            limit: Maximum number of items to return
            
        Returns:
            List of items in the specified category
        """
        self.cursor.execute("""
            SELECT oi.item_name, oi.quantity, oi.unit_price, oi.total_price,
                   o.order_id, o.order_date, oi.category
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.order_id
            WHERE oi.category = ?
            ORDER BY o.order_date DESC, oi.id DESC
            LIMIT ?
        """, (category, limit))
        
        return [
            {
                'name': row[0],
                'quantity': row[1],
                'unit_price': row[2],
                'total_price': row[3],
                'order_id': row[4],
                'order_date': row[5],
                'category': row[6]
            }
            for row in self.cursor.fetchall()
        ]
    
    def update_all_categories(self):
        """Update categories for all existing items in the database."""
        # Get all items
        self.cursor.execute("SELECT id, item_name FROM order_items")
        items = self.cursor.fetchall()
        
        # Update each item's category
        for item_id, item_name in items:
            category = self.categorize_item(item_name)
            self.cursor.execute(
                "UPDATE order_items SET category = ? WHERE id = ?",
                (category, item_id)
            )
        
        self.conn.commit()
        return len(items)
    
    def close(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()


class WalmartPDFParser:
    """Parser for Walmart order PDFs and XLSX files."""
    
    @staticmethod
    def extract_text_from_pdf(pdf_path: str) -> str:
        """Extract text content from a PDF file, using OCR if needed."""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                
                # If no text was extracted, try OCR
                if len(text.strip()) < 50:
                    try:
                        from pdf2image import convert_from_path
                        import pytesseract
                        
                        # Convert PDF to images
                        images = convert_from_path(pdf_path)
                        
                        # Extract text from each image using OCR
                        ocr_text = ""
                        for i, image in enumerate(images):
                            page_text = pytesseract.image_to_string(image)
                            ocr_text += page_text + "\n"
                        
                        if len(ocr_text.strip()) > len(text.strip()):
                            text = ocr_text
                    except Exception as ocr_error:
                        print(f"OCR extraction failed: {ocr_error}")
                
                return text
        except Exception as e:
            raise Exception(f"Failed to read PDF: {str(e)}")
    
    @staticmethod
    def parse_order_text(text: str, pdf_filename: str = "") -> Dict[str, Any]:
        """Parse Walmart order text to extract structured data."""
        order_data = {
            'order_id': None,
            'order_date': None,
            'total_amount': None,
            'subtotal': None,
            'tax': None,
            'shipping': None,
            'pdf_filename': pdf_filename,
            'raw_text': text
        }
        
        # Extract order ID (various formats)
        order_id_patterns = [
            r'Order\s*#?\s*:?\s*([0-9-]+)',
            r'Order\s+Number\s*:?\s*([0-9-]+)',
            r'#\s*([0-9]{6,}-[0-9]+)',
        ]
        for pattern in order_id_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                order_data['order_id'] = match.group(1)
                break
        
        # If no order ID found in text, try to extract from filename
        if not order_data['order_id'] and pdf_filename:
            filename_match = re.search(r'(\d{6,}-\d+)', pdf_filename)
            if filename_match:
                order_data['order_id'] = filename_match.group(1)
        
        # Extract order date
        date_patterns = [
            r'([A-Za-z]+\s+\d{1,2},\s*\d{4})\s+order',  # "Jan 11, 2026 order" format
            r'Order\s+Date\s*:?\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})',
            r'Placed\s+on\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})',
            r'(\d{1,2}/\d{1,2}/\d{4})',
        ]
        for pattern in date_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                date_str = match.group(1)
                # Convert to ISO format (YYYY-MM-DD) for proper sorting
                try:
                    from dateutil import parser as date_parser
                    parsed_date = date_parser.parse(date_str)
                    order_data['order_date'] = parsed_date.strftime('%Y-%m-%d')
                except:
                    # If parsing fails, keep original
                    order_data['order_date'] = date_str
                break
        
        # Extract total amount
        total_patterns = [
            r'Order\s+Total\s*:?\s*\$?([\d,]+\.?\d{0,2})',
            r'Total\s*:?\s*\$?([\d,]+\.?\d{0,2})',
            r'Grand\s+Total\s*:?\s*\$?([\d,]+\.?\d{0,2})',
        ]
        for pattern in total_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                order_data['total_amount'] = float(match.group(1).replace(',', ''))
                break
        
        # Extract subtotal
        subtotal_match = re.search(r'Subtotal\s*:?\s*\$?([\d,]+\.?\d{0,2})', text, re.IGNORECASE)
        if subtotal_match:
            order_data['subtotal'] = float(subtotal_match.group(1).replace(',', ''))
        
        # Extract tax
        tax_match = re.search(r'(?:Tax|Sales\s+Tax)\s*:?\s*\$?([\d,]+\.?\d{0,2})', text, re.IGNORECASE)
        if tax_match:
            order_data['tax'] = float(tax_match.group(1).replace(',', ''))
        
        # Extract shipping
        shipping_patterns = [
            r'Shipping\s*:?\s*\$?([\d,]+\.?\d{0,2})',
            r'Delivery\s*:?\s*\$?([\d,]+\.?\d{0,2})',
        ]
        for pattern in shipping_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                order_data['shipping'] = float(match.group(1).replace(',', ''))
                break
        
        return order_data
    
    @staticmethod
    def parse_items(text: str) -> List[Dict[str, Any]]:
        """Extract items from order text."""
        items = []
        
        # Walmart PDFs have items in format: [Item Name] [Status] Qty [N] $[Price]
        # Pattern to match Walmart order items
        item_pattern = r'(.+?)\s+(?:Shopped|Unavailable|Substitutions)?\s*Qty\s+(\d+)\s+\$?([\d,]+\.?\d{2})'
        
        lines = text.split('\n')
        for line in lines:
            # Skip lines that are clearly headers or totals
            if any(word in line.lower() for word in ['subtotal', 'total', 'savings', 'tax', 'delivery', 'tip', 'charge', 'payment', 'order#', 'temporary']):
                continue
            
            match = re.search(item_pattern, line)
            if match:
                name = match.group(1).strip()
                # Clean up the name - remove extra status words
                name = re.sub(r'\s+(Shopped|Unavailable|Substitutions|Weight-adjusted)$', '', name)
                
                # Skip if name is too short or looks invalid
                if len(name) < 3:
                    continue
                
                quantity = int(match.group(2))
                total_price = float(match.group(3).replace(',', ''))
                unit_price = total_price / quantity if quantity > 0 else total_price
                
                items.append({
                    'name': name[:200],  # Limit name length
                    'quantity': quantity,
                    'total_price': round(total_price, 2),
                    'unit_price': round(unit_price, 2)
                })
        
        return items
    
    @staticmethod
    def parse_xlsx(xlsx_path: str) -> tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """Parse Walmart order XLSX file.
        
        Args:
            xlsx_path: Path to the XLSX file
            
        Returns:
            Tuple of (order_data, items)
        """
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            
            # Extract order ID from filename
            filename = Path(xlsx_path).name
            order_id_match = re.search(r'Order_(.+)\.xlsx', filename)
            order_id = order_id_match.group(1) if order_id_match else None
            
            # Initialize order data
            order_data = {
                'order_id': order_id,
                'order_date': None,
                'total_amount': None,
                'subtotal': None,
                'tax': None,
                'shipping': None,
                'pdf_filename': filename,
                'raw_text': ''
            }
            
            # Parse items and totals
            items = []
            
            # Skip header row (row 1)
            for row_idx in range(2, ws.max_row + 1):
                row = list(ws[row_idx])
                product_name = row[0].value
                
                # Check if this is a total/summary row
                if product_name and isinstance(product_name, str):
                    product_name_lower = product_name.lower()
                    
                    # Extract order date
                    if 'order date' in product_name_lower:
                        date_value = str(row[1].value) if row[1].value else None
                        if date_value:
                            # Clean up date - remove extra text like "purchase"
                            date_value = re.sub(r'\s+(purchase|order|placed).*$', '', date_value, flags=re.IGNORECASE)
                            date_value = date_value.strip()
                            # Convert to ISO format (YYYY-MM-DD) for proper sorting
                            try:
                                from dateutil import parser as date_parser
                                parsed_date = date_parser.parse(date_value)
                                order_data['order_date'] = parsed_date.strftime('%Y-%m-%d')
                            except:
                                # If parsing fails, keep original
                                order_data['order_date'] = date_value
                    # Extract order number (in case it's different from filename)
                    elif 'order number' in product_name_lower:
                        if row[1].value and not order_data['order_id']:
                            order_data['order_id'] = str(row[1].value)
                    # Extract totals
                    elif 'subtotal' in product_name_lower:
                        order_data['subtotal'] = float(row[1].value) if row[1].value else 0.0
                    elif 'tax' in product_name_lower:
                        order_data['tax'] = float(row[1].value) if row[1].value else 0.0
                    elif 'delivery' in product_name_lower or 'shipping' in product_name_lower:
                        order_data['shipping'] = float(row[1].value) if row[1].value else 0.0
                    elif 'order total' in product_name_lower:
                        order_data['total_amount'] = float(row[1].value) if row[1].value else 0.0
                    elif product_name_lower in ['tip', 'delivery charges']:
                        # Skip these, just continue
                        continue
                    else:
                        # Regular item
                        quantity = row[1].value if row[1].value else 1
                        price = row[2].value if row[2].value else 0.0
                        
                        if isinstance(quantity, (int, float)) and isinstance(price, (int, float)):
                            # If price is the total price, calculate unit price
                            total_price = float(price)
                            unit_price = total_price / float(quantity) if quantity > 0 else total_price
                            
                            items.append({
                                'name': str(product_name)[:200],
                                'quantity': int(quantity),
                                'total_price': round(total_price, 2),
                                'unit_price': round(unit_price, 2)
                            })
            
            # Try to extract date from creation time if not found
            if not order_data['order_date'] and hasattr(wb.properties, 'created'):
                if wb.properties.created:
                    order_data['order_date'] = wb.properties.created.strftime('%Y-%m-%d')
            
            return order_data, items
            
        except Exception as e:
            raise Exception(f"Failed to parse XLSX file: {str(e)}")

    @staticmethod
    def parse_multirow_xlsx(xlsx_path: str) -> List[tuple[Dict[str, Any], List[Dict[str, Any]]]]:
        """Parse Walmart orders XLSX file with one row per item format.
        
        Supports two formats:
        Format 1: Order Number, Order Date, Subtotal, Order Total, Product Name, Quantity, Price, Delivery Status
        Format 2: Order Number, Order Date, Shipping Address, Payment Method, Subtotal, Order Total, Product Name, Quantity, Price
        
        Args:
            xlsx_path: Path to the XLSX file
            
        Returns:
            List of tuples of (order_data, items) for each unique order
        """
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            
            # Detect format by checking headers
            first_row = [cell.value for cell in ws[1]]
            
            # Determine column indices based on format
            if len(first_row) > 6 and first_row[6] == 'Product Name':
                # Format 2: Has Shipping Address and Payment Method
                col_indices = {
                    'order_id': 0,
                    'order_date': 1,
                    'subtotal': 4,
                    'order_total': 5,
                    'product_name': 6,
                    'quantity': 7,
                    'price': 8,
                    'delivery_status': 12 if len(first_row) > 12 else None
                }
            else:
                # Format 1: Direct to Subtotal
                col_indices = {
                    'order_id': 0,
                    'order_date': 1,
                    'subtotal': 2,
                    'order_total': 3,
                    'product_name': 4,
                    'quantity': 5,
                    'price': 6,
                    'delivery_status': 7 if len(first_row) > 7 else None
                }
            
            orders_dict = {}  # Dictionary to group items by order_id
            
            # Process each row (skip header row 1)
            for row_idx in range(2, ws.max_row + 1):
                row = list(ws[row_idx])
                
                # Extract values from columns using detected indices
                order_id = str(row[col_indices['order_id']].value) if row[col_indices['order_id']].value else None
                order_date_str = str(row[col_indices['order_date']].value) if row[col_indices['order_date']].value else None
                subtotal = float(row[col_indices['subtotal']].value) if row[col_indices['subtotal']].value and isinstance(row[col_indices['subtotal']].value, (int, float)) else 0.0
                order_total = float(row[col_indices['order_total']].value) if row[col_indices['order_total']].value and isinstance(row[col_indices['order_total']].value, (int, float)) else 0.0
                product_name = str(row[col_indices['product_name']].value) if row[col_indices['product_name']].value else None
                quantity = int(row[col_indices['quantity']].value) if row[col_indices['quantity']].value and isinstance(row[col_indices['quantity']].value, (int, float)) else 1
                price = float(row[col_indices['price']].value) if row[col_indices['price']].value and isinstance(row[col_indices['price']].value, (int, float)) else 0.0
                
                # Extract delivery status if available
                delivery_status = None
                if col_indices.get('delivery_status') is not None and len(row) > col_indices['delivery_status']:
                    delivery_status_val = row[col_indices['delivery_status']].value
                    if delivery_status_val:
                        delivery_status = str(delivery_status_val)
                
                if not order_id or not product_name:
                    continue
                
                # Convert date to ISO format
                order_date = None
                if order_date_str:
                    try:
                        from dateutil import parser as date_parser
                        parsed_date = date_parser.parse(order_date_str)
                        order_date = parsed_date.strftime('%Y-%m-%d')
                    except:
                        order_date = order_date_str
                
                # Create or update order data
                if order_id not in orders_dict:
                    orders_dict[order_id] = {
                        'order_data': {
                            'order_id': order_id,
                            'order_date': order_date,
                            'total_amount': order_total,
                            'subtotal': subtotal if subtotal > 0 else None,
                            'tax': None,  # Not provided in this format
                            'shipping': None,  # Not provided in this format
                            'pdf_filename': Path(xlsx_path).name,
                            'raw_text': ''
                        },
                        'items': []
                    }
                
                # Add item with delivery status
                unit_price = price / quantity if quantity > 0 else price
                orders_dict[order_id]['items'].append({
                    'name': product_name[:200],
                    'quantity': quantity,
                    'total_price': round(price, 2),
                    'unit_price': round(unit_price, 2),
                    'delivery_status': delivery_status
                })
            
            # Return list of (order_data, items) tuples
            return [(order['order_data'], order['items']) for order in orders_dict.values()]
            
        except Exception as e:
            raise Exception(f"Failed to parse multi-row XLSX file: {str(e)}")


async def execute(pdf_path: Optional[str] = None, action: str = "parse") -> Dict[str, Any]:
    """
    Execute Walmart order operations.
    
    Args:
        pdf_path: Path to PDF file or None to process all PDFs
        action: Action to perform - "parse", "list", "query", "stats", or "items"
    
    Returns:
        Dictionary with results
    """
    db = WalmartOrderDB()
    parser = WalmartPDFParser()
    
    try:
        if action == "parse":
            if pdf_path:
                # Parse single file (PDF or XLSX)
                full_path = Path(pdf_path)
                if not full_path.is_absolute():
                    full_path = Path("data/walmart") / pdf_path
                
                if not full_path.exists():
                    return {"error": f"File not found: {pdf_path}"}
                
                # Check file extension
                if full_path.suffix.lower() == '.xlsx':
                    # Parse XLSX
                    order_data, items = parser.parse_xlsx(str(full_path))
                else:
                    # Parse PDF
                    text = parser.extract_text_from_pdf(str(full_path))
                    order_data = parser.parse_order_text(text, full_path.name)
                    items = parser.parse_items(text)
                
                if not order_data['order_id']:
                    return {"error": f"Could not extract order ID from {full_path.suffix.upper()}"}
                
                db.insert_order(order_data, items)
                
                return {
                    "success": True,
                    "message": f"Parsed order {order_data['order_id']} with {len(items)} items",
                    "order": order_data,
                    "items": items,
                    "items_count": len(items)
                }
            else:
                # Parse all PDFs and XLSX files in walmart folder and subfolders
                walmart_dir = Path("data/walmart")
                pdf_files = list(walmart_dir.glob("*.pdf")) + list(walmart_dir.glob("archived/*.pdf"))
                xlsx_files = list(walmart_dir.glob("**/*.xlsx"))
                all_files = pdf_files + xlsx_files
                
                results = []
                for file_path in all_files:
                    try:
                        if file_path.suffix.lower() == '.xlsx':
                            # Check if it's a multi-row format file
                            if file_path.name == 'Walmart_Orders.xlsx':
                                # Parse multi-row XLSX
                                orders_list = parser.parse_multirow_xlsx(str(file_path))
                                for order_data, items in orders_list:
                                    db.insert_order(order_data, items)
                                results.append({
                                    "file": file_path.name,
                                    "order_id": f"{len(orders_list)} orders",
                                    "items_count": sum(len(items) for _, items in orders_list),
                                    "success": True
                                })
                            else:
                                # Parse single-order XLSX
                                order_data, items = parser.parse_xlsx(str(file_path))
                                if order_data['order_id']:
                                    db.insert_order(order_data, items)
                                    results.append({
                                        "file": file_path.name,
                                        "order_id": order_data['order_id'],
                                        "items_count": len(items),
                                        "success": True
                                    })
                                else:
                                    results.append({
                                        "file": file_path.name,
                                        "success": False,
                                        "error": "Could not extract order ID"
                                    })
                        else:
                            # Parse PDF
                            text = parser.extract_text_from_pdf(str(file_path))
                            order_data = parser.parse_order_text(text, file_path.name)
                            items = parser.parse_items(text)
                        
                            if order_data['order_id']:
                                db.insert_order(order_data, items)
                                results.append({
                                    "file": file_path.name,
                                    "order_id": order_data['order_id'],
                                    "items_count": len(items),
                                    "success": True
                                })
                            else:
                                results.append({
                                    "file": file_path.name,
                                    "success": False,
                                    "error": "Could not extract order ID"
                                })
                    except Exception as e:
                        results.append({
                            "file": file_path.name,
                            "success": False,
                            "error": str(e)
                        })
                
                return {
                    "success": True,
                    "message": f"Processed {len(all_files)} files ({len(pdf_files)} PDFs, {len(xlsx_files)} XLSX)",
                    "results": results
                }
        
        elif action == "list":
            orders = db.get_all_orders(limit=50)
            return {
                "success": True,
                "orders": orders,
                "count": len(orders)
            }
        
        elif action == "stats":
            total_spent = db.get_total_spent()
            all_orders = db.get_all_orders(limit=1000)
            
            return {
                "success": True,
                "total_spent": total_spent,
                "total_orders": len(all_orders),
                "average_order": total_spent / len(all_orders) if all_orders else 0
            }
        
        elif action == "query":
            if pdf_path:  # Use pdf_path as order_id for query
                order = db.get_order(pdf_path)
                if order:
                    return {
                        "success": True,
                        "order": order
                    }
                else:
                    return {"error": f"Order {pdf_path} not found"}
            else:
                return {"error": "No order ID provided for query"}
        
        elif action == "items":
            # List all items or search
            if pdf_path:  # Use as search term
                items = db.search_items(pdf_path)
                return {
                    "success": True,
                    "items": items,
                    "search_term": pdf_path
                }
            else:
                items = db.get_all_items(limit=100)
                return {
                    "success": True,
                    "items": items,
                    "count": len(items)
                }
        
        else:
            return {"error": f"Unknown action: {action}"}
    
    finally:
        db.close()
