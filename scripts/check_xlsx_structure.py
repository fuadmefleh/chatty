#!/usr/bin/env python3
import openpyxl

xlsx_path = '/home/edgeworks-server/chatty/data/walmart/Walmart_Orders (3).xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# Print headers
first_row = [cell.value for cell in ws[1]]
print(f"Total columns: {len(first_row)}\n")
print("Headers:")
for i, header in enumerate(first_row):
    print(f"  Column {i}: {header}")

# Print first data row
print("\nFirst item row:")
second_row = [cell.value for cell in ws[2]]
for i, value in enumerate(second_row):
    if value is not None:
        print(f"  Column {i}: {value}")
